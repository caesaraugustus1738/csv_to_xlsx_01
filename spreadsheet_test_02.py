import csv
import shutil
import openpyxl
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Font, Side

COLOUR_INDEX = (
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
    '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
    '00993300', '00993366', '00333399', '00333333',  #60-63
)

COLOUR_INDEX_sliced = COLOUR_INDEX[41:48]

'''
Put CSV into dataframe
Sort columns
Write to XLSX file
Style XLSX file
'''

# --- Make base XLSX ---

# Get raw CSV
source_file = Path(sys.argv[1])

print(source_file.name)

# Define a dataframe
csv_df = pd.read_csv(source_file)

# Setting so dataframe.head shows all columns
pd.set_option('display.max_columns',None)

# # Print all columns in dataframe
# print(csv_df.head)

# Sort a series without affecting dataframe
# print(csv_data_mod['lens'].sort_values(ascending=False))

# # Specific dataframe column
# print(csv_df['source timecode in'])

# Sort df by two indexes (columns)
csv_df = csv_df.sort_values(['work in progress notes','Shot Number'])

# Write dataframe to file
with pd.ExcelWriter('shot_export_001_fmttd.xlsx') as writer:
	csv_df.to_excel(writer)

# --- Format XLSX ---

# turq = PatternFill(start_color = '00CCFFFF', fill_type='solid')
# blue = PatternFill(start_color = COLOUR_INDEX)
font = Font(bold = False)
side = Side(border_style = None)
no_borders = Border(left=side,right=side,top=side,bottom=side)

# Load workbook (xlsx)
wb = load_workbook(writer)

# Specify sheetname
ws = wb[wb.sheetnames[0]]

# Remove borders and bold font
for row in ws:
	for cell in row:
		cell.border = no_borders
		cell.font = font

# Hide first column
ws.column_dimensions['A'].hidden = True

for i in range(41,48):
	ws['K' + str(i+1)].fill = PatternFill(start_color = COLOUR_INDEX[i], fill_type = 'solid')
	ws['L' + str(i+1)] = i+1
	# print('K'+str(i))

for cell in ws[1][1:]:
	if cell.value:
		cell.fill = PatternFill(start_color = COLOUR_INDEX[40 + ws[1].index(cell)], fill_type='solid')

wb.save(str(source_file.name).replace('.csv','.xlsx'))


# # --- Make a list from CSV data ---

# with open(source_file, encoding = 'utf-8-sig') as file:
# 	# Create reader object
# 	reader = csv.reader(file)
	
# 	# Make reader object a list
# 	reader_list = list(reader)

# 	# print('Hey!', reader_list)

# 	# Add CSV headings to a list
# 	csv_list = []
# 	for i in reader_list[0]:
# 		csv_list.append([i])

# 	# Add entries to same list as heading
# 	for i in reader_list[1:]:
# 		for j in range(len(reader_list[0])):
# 			csv_list[j].append(i[j])

# 	# Make strings integers
# 	for i in range(len(csv_list)):
# 		for j in range(len(csv_list[i])):
# 			try:
# 				csv_list[i][j] = int(csv_list[i][j])
# 			except:
# 				pass

# print(csv_list)

# # -- Write csv list into xlsx --

# # Set xlsx title
# xlsx_title = source_file[:-4]+'_fmttd.xlsx'

# # Create workbook
# wb = Workbook()

# # Save workbook to create on disk
# wb.save(xlsx_title)

# # Load existing workbook
# wb = load_workbook(xlsx_title)

# # Access specific worksheet
# ws = wb[wb.sheetnames[0]]

# for i in range(len(csv_list)):
# 	for j in range(len(csv_list[0])):
# 		ws.cell(row = j+1, column = i+1, value = csv_list[i][j])

# ws['A1'].fill = fl

# wb.save(xlsx_title)

